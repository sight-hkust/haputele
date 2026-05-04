"""Bulk daily exports for the healthworker:

- `GET /exports/medications.xlsx` — pickup spreadsheet (one row per medication)
- `GET /exports/prescriptions.zip` — every signed prescription PDF as a zip

Both endpoints take a `from` / `to` ISO datetime window and return only
`completed` appointments (the only state with a finalised consultation + meds).

Day boundaries, filename date stamps, and the patient-age column are
anchored to the export timezone from `system_config.export_timezone`
(seeded at /setup/initialize, defaulting to Asia/Colombo for SL
pharmacies). The client is expected to use the same zone when computing
the `from` / `to` window (see `frontend/src/lib/format.ts:EXPORT_TIMEZONE`).
"""
from datetime import datetime, timezone
from io import BytesIO
import re
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from ..deps import db_dep, require_role
from ..models import Appointment, Consultation, Doctor, Patient
from ..pdf import render_prescription_pdf
from ..tz import export_tz


router = APIRouter(prefix="/exports", tags=["exports"])


_HEADERS = [
    "Appointment", "Patient ID", "Patient Name", "National ID", "Gender",
    "Age", "Contact", "Doctor", "SLMC", "Generic Name", "Trade Name",
    "Dose", "Frequency", "Duration", "Instructions",
]


def _completed_in_window(db: Session, frm: datetime, to: datetime):
    stmt = (
        select(Appointment, Patient, Doctor, Consultation)
        .join(Patient, Patient.patient_id == Appointment.patient_id)
        .join(Doctor, Doctor.doctor_id == Appointment.doctor_id)
        .join(Consultation, Consultation.appointment_id == Appointment.appointment_id)
        .where(
            and_(
                Appointment.scheduled_at >= frm,
                Appointment.scheduled_at <= to,
                Appointment.status == "completed",
                Consultation.status == "completed",
            )
        )
        .order_by(Appointment.scheduled_at)
    )
    return db.execute(stmt).all()


def _safe_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", s).strip("_") or "patient"


def _age_from_dob(dob) -> int | None:
    if not dob:
        return None
    today = datetime.now(timezone.utc).astimezone(export_tz()).date()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


@router.get(
    "/medications.xlsx",
    dependencies=[Depends(require_role("healthworker", "admin"))],
)
def export_medications_xlsx(
    frm: datetime = Query(..., alias="from"),
    to: datetime = Query(...),
    db: Session = Depends(db_dep),
) -> Response:
    rows = _completed_in_window(db, frm, to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pickup list"

    # Title block
    ws["A1"] = f"Medication Pickup — {frm.astimezone(export_tz()).date().isoformat()}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))

    total_patients = len({r[0].patient_id for r in rows})
    total_meds = sum(len((r[3].medications or [])) for r in rows)
    ws["A2"] = (
        f"{len(rows)} appointment(s) · {total_patients} patient(s) · {total_meds} medication(s)"
    )
    ws["A2"].font = Font(italic=True, color="64748B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_HEADERS))

    # Header row
    HEADER_ROW = 4
    for col, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data
    r = HEADER_ROW + 1
    for appt, pat, doc, consult in rows:
        meds = consult.medications or []
        base = [
            appt.appointment_id,
            pat.patient_id,
            f"{pat.given_name} {pat.family_name}",
            pat.n_id or "",
            pat.gender,
            _age_from_dob(pat.dob),
            pat.contact or "",
            f"Dr. {doc.given_name} {doc.family_name}",
            doc.slmc_registration_number,
        ]
        if not meds:
            ws.append(base + ["(no medications)", "", "", "", "", ""])
            r += 1
        else:
            for m in meds:
                ws.append(
                    base
                    + [
                        m.get("genericName", ""),
                        m.get("tradeName", ""),
                        m.get("dose", ""),
                        m.get("frequency", ""),
                        m.get("duration", ""),
                        m.get("instructions", ""),
                    ]
                )
                r += 1

    # Auto column widths (capped — long instruction strings shouldn't blow the layout)
    for col_idx in range(1, len(_HEADERS) + 1):
        col_letter = ws.cell(row=HEADER_ROW, column=col_idx).column_letter
        max_len = 10
        for row_idx in range(HEADER_ROW, max(r, HEADER_ROW + 1)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="medication-pickup-{frm.astimezone(export_tz()).date().isoformat()}.xlsx"'
            )
        },
    )


@router.get(
    "/prescriptions.zip",
    dependencies=[Depends(require_role("healthworker", "admin"))],
)
def export_prescriptions_zip(
    frm: datetime = Query(..., alias="from"),
    to: datetime = Query(...),
    db: Session = Depends(db_dep),
) -> Response:
    rows = _completed_in_window(db, frm, to)

    buf = BytesIO()
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        manifest = [
            f"Prescription PDFs — {frm.astimezone(export_tz()).date().isoformat()}",
            f"{len(rows)} prescription(s)",
            "",
        ]
        for appt, pat, doc, consult in rows:
            pdf_bytes = render_prescription_pdf(
                patient=pat, doctor=doc, appointment=appt, consultation=consult,
            )
            patient_slug = _safe_filename(f"{pat.given_name}_{pat.family_name}")
            filename = f"Rx-{appt.appointment_id}-{patient_slug}.pdf"
            zf.writestr(filename, pdf_bytes)
            manifest.append(
                f"{filename}  ·  {pat.given_name} {pat.family_name}  ·  Dr. {doc.given_name} {doc.family_name}"
            )
        zf.writestr("manifest.txt", "\n".join(manifest))

    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f'attachment; filename="prescriptions-{frm.astimezone(export_tz()).date().isoformat()}.zip"'
            )
        },
    )
